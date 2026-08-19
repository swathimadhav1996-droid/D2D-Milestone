"""
OD2D Milestone Completeness — Streamlit app
-------------------------------------------
Upload the raw shipment-level export (the CSV produced by the OD2D shipment-level
SQL query) and download a formatted "Completeness Summary" workbook: overall
completeness + a breakdown by FFW / NVOCC / Carrier, one column per milestone,
with green/yellow/red traffic-light styling.

Sidebar filters:
  - Shipments to include (all rows / completed only), optional created-date range
  - Milestones to include — pick exactly which of the 24 milestones appear as
    columns in the report (Select all / Clear all shortcuts provided)

Shipment identity: every row in the file is counted as its own shipment — both
booking-level (parent) and container-level (child) rows are always combined,
and rows are NEVER merged or dropped just because they share the same
CONTAINER_ID or the same BOOKING_NUMBER + CONTAINER_ID pair. Container IDs get
reused across different bookings over time, so CONTAINER_ID alone is never a
reliable identity — the app surfaces a Booking+Container duplicate check (see
the "Data quality check" panel) purely for visibility, it does not remove rows.

Percentage methodology (read this if a manual check doesn't match): every
percentage — including each FFW subtotal and the "Overall Completeness" row
at the top — is a SHIPMENT-WEIGHTED average, i.e. (shipments where that
milestone is present) / (shipments in that row's scope). It is NOT a plain
average of the rows below it. Selecting a spreadsheet range that spans both
the subtotal rows AND the individual entity rows above them (e.g. Excel's
AVERAGE() over a whole block) will double-count and will not reproduce the
top total — weight by each row's Shipment Count instead (SUMPRODUCT of
count x percentage, divided by SUM of count) to verify by hand.

Optionally includes ONE raw-data sheet in the download, with every column from
the uploaded file, matching whichever "Shipments to include" scope is
selected — "Raw Data - All" when All rows is selected, or "Raw Data -
Completed" when Completed only is selected. Never both at once.

Run locally:   streamlit run streamlit_app.py
Deploy:        push this file + requirements.txt to GitHub, then deploy on
               https://share.streamlit.io  (main file path = streamlit_app.py)
"""

import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Milestone definitions — one column per individual milestone, in journey order.
# (name shown in the sheet, list of timestamp columns, journey category;
#  "present" = any of the listed timestamp columns is filled)
# ---------------------------------------------------------------------------
ALL_GROUPS = [
    ("Gate Out Empty At Terminal",            ["TS_GATE_OUT_EMPTY_AT_TERMINAL"],            "Empty Pickup"),
    # --- Pre Carriage ---
    ("Picked Up At Origin",                   ["TS_PICKED_UP_AT_ORIGIN"],                   "Pre Carriage"),
    ("Arrival At Origin CFS/WH",              ["TS_ARRIVAL_ORIGIN_CFS_OR_WH"],              "Pre Carriage"),
    ("Load At Origin CFS/WH",                 ["TS_LOAD_ORIGIN_CFS_OR_WH"],                 "Pre Carriage"),
    ("Arrival At Inland Export Terminal",     ["TS_ARRIVAL_INLAND_EXPORT_TERMINAL"],        "Pre Carriage"),
    ("Discharge At Inland Export Terminal",   ["TS_DISCHARGE_INLAND_EXPORT_TERMINAL"],      "Pre Carriage"),
    ("Load At Inland Export Terminal",        ["TS_LOAD_INLAND_EXPORT_TERMINAL"],           "Pre Carriage"),
    ("Departure From Inland Export Terminal", ["TS_DEPARTURE_INLAND_EXPORT_TERMINAL"],      "Pre Carriage"),
    # --- Port to port ---
    ("Gate In Full At POL",                   ["TS_GATE_IN_FULL_POL"],                      "Port to Port"),
    ("Load Onto Vessel At POL",               ["TS_LOAD_ONTO_VESSEL_POL"],                  "Port to Port"),
    ("Vessel Departure From POL",             ["TS_VESSEL_DEPARTURE_POL"],                  "Port to Port"),
    ("Vessel Arrival At POD",                 ["TS_VESSEL_ARRIVAL_POD"],                    "Port to Port"),
    ("Discharge From Vessel At POD",          ["TS_DISCHARGE_FROM_VESSEL_POD"],             "Port to Port"),
    ("Gate Out Full At POD",                  ["TS_GATE_OUT_FULL_POD"],                     "Port to Port"),
    # --- On Carriage ---
    ("Arrival (Rail) Inland Import Terminal", ["TS_ARRIVAL_RAIL_INLAND_IMPORT"],             "On Carriage"),
    ("Arrival At Inland Import Terminal",     ["TS_ARRIVAL_INLAND_IMPORT_TERMINAL"],        "On Carriage"),
    ("Discharge At Inland Import Terminal",   ["TS_DISCHARGE_INLAND_IMPORT_TERMINAL"],      "On Carriage"),
    ("Load At Inland Import Terminal",        ["TS_LOAD_INLAND_IMPORT_TERMINAL"],           "On Carriage"),
    ("Departure From Inland Import Terminal", ["TS_DEPARTURE_INLAND_IMPORT_TERMINAL"],      "On Carriage"),
    ("Out For Delivery",                      ["TS_OUT_FOR_DELIVERY"],                      "On Carriage"),
    ("Arrival Of Full Container At Consignee",["TS_ARRIVAL_AT_CONSIGNEE"],                  "On Carriage"),
    ("Proof Of Delivery",                     ["TS_PROOF_OF_DELIVERY"],                     "On Carriage"),
    ("Picked Up Empty From Consignee",        ["TS_EMPTY_PICKUP_FROM_CONSIGNEE"],           "On Carriage"),
    # --- Empty Return ---
    ("Gate In Empty At Terminal",             ["TS_GATE_IN_EMPTY_AT_TERMINAL"],             "Empty Return"),
]
ALL_GNAMES = [g for g, _, _ in ALL_GROUPS]
MILESTONE_CATEGORY = {g: cat for g, _, cat in ALL_GROUPS}

FFW_COL, NVOCC_COL, CARRIER_COL = "MASTER_FFW_NAME", "MASTER_NVOCC_NAME", "MASTER_CARRIER_NAME"

# colours (match the reference workbook)
NAVY, BLUE, F_LBL, F_CNT = "1F4E79", "2E75B6", "D6E4F7", "D9E1F2"
GREEN_F, GREEN_T = "C6EFCE", "276221"
YEL_F, YEL_T     = "FFEB9C", "7D6608"
RED_F, RED_T     = "FFC7CE", "9C0006"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(THIN, THIN, THIN, THIN)


def _nonempty(series: pd.Series) -> pd.Series:
    s = series.astype("string")
    return s.notna() & (s.str.strip() != "") & (s.str.lower() != "nan")


def compute(df: pd.DataFrame, groups=None):
    """Return (overall dict, entity DataFrame, ffw-subtotal DataFrame, N, pcols).

    `groups` is a list of (name, [timestamp_columns], category) tuples — pass a
    subset of ALL_GROUPS to restrict the report to only those milestones.
    Defaults to every milestone (ALL_GROUPS) if not given.
    """
    groups = groups if groups is not None else ALL_GROUPS
    gnames = [g for g, _, _ in groups]
    df = df.copy()
    # ensure every needed timestamp column exists
    for _, cols, _ in groups:
        for c in cols:
            if c not in df.columns:
                df[c] = pd.NA
    # present flag per milestone
    for g, cols, _ in groups:
        flags = pd.concat([_nonempty(df[c]) for c in cols], axis=1)
        df["_" + g] = flags.any(axis=1).astype(int)
    # clean party names
    for c in [FFW_COL, NVOCC_COL, CARRIER_COL]:
        if c not in df.columns:
            df[c] = pd.NA
        df[c] = df[c].where(_nonempty(df[c]), other=None)

    pcols = ["_" + g for g in gnames]
    N = len(df)
    overall = {g: df["_" + g].mean() for g in gnames}

    key_ffw = df[FFW_COL].fillna("—")
    ent = (df.groupby([key_ffw, df[NVOCC_COL].fillna("—"), df[CARRIER_COL].fillna("—")], dropna=False)
             .agg(cnt=(FFW_COL, "size"), **{p: (p, "mean") for p in pcols})
             .reset_index())
    ent.columns = ["FFW", "NVOCC", "CARRIER", "cnt"] + pcols

    fsub = (df.groupby(key_ffw).agg(cnt=(FFW_COL, "size"), **{p: (p, "mean") for p in pcols}).reset_index())
    fsub.columns = ["FFW", "cnt"] + pcols
    return overall, ent, fsub, N, pcols


def band(v):
    if v >= 0.80: return GREEN_F, GREEN_T
    if v >= 0.50: return YEL_F, YEL_T
    return RED_F, RED_T


def build_workbook(overall, ent, fsub, N, pcols, gnames) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Completeness Summary"
    NC = 4 + len(gnames)

    def pct(cell, v, fill=True, bold=False):
        cell.value = v; cell.number_format = "0%"
        f, t = band(v)
        cell.font = Font(name="Calibri", bold=bold, color=t)
        if fill: cell.fill = PatternFill("solid", fgColor=f)
        cell.alignment = Alignment(horizontal="center"); cell.border = BORDER

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, f"Overall Completeness  |  Total Records: {N:,}")
    c.font = Font(name="Calibri", bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    r += 1
    for i, g in enumerate(gnames):
        cc = ws.cell(r, i + 1, g); cc.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor=BLUE); cc.alignment = Alignment(horizontal="center", wrap_text=True); cc.border = BORDER
    r += 1
    for i, g in enumerate(gnames):
        cc = ws.cell(r, i + 1, overall[g]); cc.number_format = "0%"
        cc.font = Font(name="Calibri", bold=True, color=YEL_T); cc.fill = PatternFill("solid", fgColor=YEL_F)
        cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
    r += 1

    # Methodology note — prevents "why doesn't AVERAGE() of the rows below match this?" confusion
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    note = ws.cell(r, 1, "Note: every % here is a shipment-weighted average (shipments with the milestone ÷ "
                         "shipments in scope) — not a plain average of the subtotal/entity rows below. "
                         "To verify a total by hand, weight each row by its Shipment Count "
                         "(SUMPRODUCT(count, %) ÷ SUM(count)), don't AVERAGE() the % column directly.")
    note.font = Font(name="Calibri", italic=True, size=9, color="7F7F7F")
    note.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Completeness by Entity (FFW / NVOCC / Carrier)")
    c.font = Font(name="Calibri", bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    r += 1
    for i, h in enumerate(["FFW Name", "NVOCC Name", "Carrier Name", "Shipment Count"] + gnames):
        cc = ws.cell(r, i + 1, h); cc.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor=BLUE); cc.alignment = Alignment(horizontal="center", wrap_text=True); cc.border = BORDER
    header_row = r
    r += 1

    def lbl(cell, val, bold=False, color="000000"):
        cell.value = val; cell.font = Font(name="Calibri", bold=bold, color=color)
        cell.fill = PatternFill("solid", fgColor=F_LBL); cell.border = BORDER; cell.alignment = Alignment(horizontal="left")

    def cntcell(cell, val, bold=True):
        cell.value = int(val); cell.number_format = "#,##0"; cell.font = Font(name="Calibri", bold=bold, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=F_CNT); cell.alignment = Alignment(horizontal="center"); cell.border = BORDER

    ffws = sorted([f for f in ent["FFW"].unique() if f != "—"])
    order = (["—"] if "—" in ent["FFW"].values else []) + ffws
    for ffw in order:
        rows = ent[ent["FFW"] == ffw].sort_values("cnt", ascending=False)
        for _, row in rows.iterrows():
            lbl(ws.cell(r, 1), row["FFW"]); lbl(ws.cell(r, 2), row["NVOCC"]); lbl(ws.cell(r, 3), row["CARRIER"])
            cntcell(ws.cell(r, 4), row["cnt"], bold=False)
            for i, p in enumerate(pcols): pct(ws.cell(r, 5 + i), row[p])
            r += 1
        s = fsub[fsub["FFW"] == ffw].iloc[0]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        lbl(ws.cell(r, 1), "Subtotal — No FFW" if ffw == "—" else f"Subtotal — {ffw}", bold=True, color=NAVY)
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=F_LBL); ws.cell(r, 3).fill = PatternFill("solid", fgColor=F_LBL)
        cntcell(ws.cell(r, 4), s["cnt"], bold=True)
        for i, p in enumerate(pcols): pct(ws.cell(r, 5 + i), s[p], bold=True)
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    tc = ws.cell(r, 1, "Total"); tc.font = Font(name="Calibri", bold=True, color="FFFFFF"); tc.fill = PatternFill("solid", fgColor=NAVY); tc.border = BORDER
    for cc in (ws.cell(r, 2), ws.cell(r, 3)): cc.fill = PatternFill("solid", fgColor=NAVY); cc.border = BORDER
    tt = ws.cell(r, 4, N); tt.number_format = "#,##0"; tt.font = Font(name="Calibri", bold=True, color="FFFFFF")
    tt.fill = PatternFill("solid", fgColor=NAVY); tt.alignment = Alignment(horizontal="center"); tt.border = BORDER
    for i, g in enumerate(gnames):
        cc = ws.cell(r, 5 + i, overall[g]); cc.number_format = "0%"; cc.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor=NAVY); cc.alignment = Alignment(horizontal="center"); cc.border = BORDER

    ws.freeze_panes = "E" + str(header_row + 1)
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 34; ws.column_dimensions["D"].width = 15
    for i in range(len(gnames)): ws.column_dimensions[get_column_letter(5 + i)].width = 15

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def add_raw_sheet(wb, name: str, data: pd.DataFrame):
    """Append a plain raw-data sheet (header + every row/column, filterable) to wb."""
    ws = wb.create_sheet(name)
    cols = list(data.columns)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(1, j, col)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(data.itertuples(index=False, name=None), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, None if pd.isna(val) else val)
    ws.freeze_panes = "A2"
    if len(data):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data) + 1}"
    sample = data.head(300)
    for j, col in enumerate(cols, start=1):
        maxlen = len(str(col))
        for v in sample.iloc[:, j - 1]:
            if pd.notna(v):
                maxlen = max(maxlen, min(len(str(v)), 40))
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(maxlen + 2, 32))
    return ws


def build_full_workbook(overall, ent, fsub, N, pcols, gnames, raw_df=None, raw_sheet_name=None) -> bytes:
    """Build the Completeness Summary workbook, optionally appending ONE raw-data sheet."""
    summary_bytes = build_workbook(overall, ent, fsub, N, pcols, gnames)
    if raw_df is None:
        return summary_bytes
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(summary_bytes))
    add_raw_sheet(wb, raw_sheet_name or "Raw Data", raw_df)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ============================== UI ==============================
st.set_page_config(page_title="OD2D Milestone Completeness", layout="wide")
st.title("OD2D Milestone Completeness Builder")
st.caption("Upload the raw shipment-level export and download the formatted completeness workbook.")

with st.sidebar:
    st.header("Options")
    st.caption("Every row is counted as its own shipment — booking-level and "
               "container-level rows are always combined, and nothing is merged "
               "or dropped for sharing a Container ID or Booking+Container pair.")
    scope = st.radio("Shipments to include",
                     ["All rows", "Completed only"], index=0)
    use_dates = st.checkbox("Filter by shipment created date", value=False)
    d_from = d_to = None
    if use_dates:
        d_from = st.date_input("Created from")
        d_to = st.date_input("Created to")

    st.markdown("---")
    st.subheader("Milestones to include")

    if "selected_milestones" not in st.session_state:
        st.session_state.selected_milestones = list(ALL_GNAMES)

    bcol1, bcol2 = st.columns(2)
    if bcol1.button("Select all", use_container_width=True):
        st.session_state.selected_milestones = list(ALL_GNAMES)
    if bcol2.button("Clear all", use_container_width=True):
        st.session_state.selected_milestones = []

    selected_milestones = st.multiselect(
        "Milestones",
        options=ALL_GNAMES,
        default=st.session_state.selected_milestones,
        key="selected_milestones",
        format_func=lambda m: f"{m}  ·  {MILESTONE_CATEGORY[m]}",
        label_visibility="collapsed",
    )
    st.caption(f"{len(selected_milestones)} of {len(ALL_GNAMES)} milestones selected.")

    st.markdown("---")
    include_raw = st.checkbox("Include a raw-data sheet in the download", value=True)
    st.caption("Adds one extra tab with every column from the uploaded file, matching whichever "
                "'Shipments to include' scope is selected above — 'Raw Data - All' when All rows is "
                "selected, or 'Raw Data - Completed' when Completed only is selected. "
                "Turn off for a smaller/faster file if you only need the summary.")

    st.markdown("---")
    st.caption("Green ≥ 80%  ·  Yellow 50–79%  ·  Red < 50%")

up = st.file_uploader("Raw shipment-level file (CSV or Excel)", type=["csv", "xlsx"])

if up is not None:
    df = pd.read_csv(up, dtype=str) if up.name.lower().endswith(".csv") else pd.read_excel(up, dtype=str)
    st.write(f"Loaded **{len(df):,}** rows — booking-level and container-level rows combined, nothing filtered by grain.")

    # ---- Data quality check: Booking + Container identity (informational only — never dedups) ----
    if "BOOKING_NUMBER" in df.columns or "CONTAINER_ID" in df.columns:
        bn = df["BOOKING_NUMBER"] if "BOOKING_NUMBER" in df.columns else pd.Series([pd.NA] * len(df), index=df.index)
        cid = df["CONTAINER_ID"] if "CONTAINER_ID" in df.columns else pd.Series([pd.NA] * len(df), index=df.index)
        shipment_key = bn.fillna("").astype(str).str.strip() + " | " + cid.fillna("").astype(str).str.strip()
        has_key = shipment_key.str.strip(" |") != ""
        dup_mask = shipment_key.duplicated(keep=False) & has_key
        n_dup_rows = int(dup_mask.sum())
        n_dup_keys = int(shipment_key[dup_mask].nunique()) if n_dup_rows else 0

        with st.expander(f"Data quality check — Booking + Container identity  ({n_dup_rows:,} rows share a repeated key)"):
            st.write(f"- Unique Booking Number + Container ID combinations: **{shipment_key[has_key].nunique():,}**")
            st.write(f"- Rows sharing a repeated Booking+Container combination: **{n_dup_rows:,}** "
                     f"across **{n_dup_keys:,}** distinct combinations")
            st.write("All rows are kept exactly as loaded — nothing is merged or dropped. A Container ID reused "
                     "under a *different* Booking Number is a different shipment (containers get reused over "
                     "time), and even a true Booking+Container repeat is preserved as its own row.")
            if n_dup_rows:
                preview_cols = [c for c in ["BOOKING_NUMBER", "CONTAINER_ID", "INTERNAL_SHIPMENT_ID", "BOOKING_LEVEL"] if c in df.columns]
                st.dataframe(df.loc[dup_mask, preview_cols].sort_values(preview_cols[:2]).head(50),
                             use_container_width=True, hide_index=True)

    # ---- Completion status (computed on the full loaded set, before the Completed-only filter) ----
    if "SHIPMENT_COMPLETED_DT" in df.columns:
        completed_n = int(_nonempty(df["SHIPMENT_COMPLETED_DT"]).sum())
        total_n = len(df)
        c1, c2, c3 = st.columns(3)
        c1.metric("Total rows loaded", f"{total_n:,}")
        c2.metric("Completed", f"{completed_n:,}", f"{completed_n / total_n:.0%}" if total_n else None)
        c3.metric("Not yet completed", f"{total_n - completed_n:,}")
        st.caption("A shipment only gets a completed date once its full door-to-door journey has finished — "
                   "shipments created recently are expected to still show as not-yet-completed.")

    if scope == "Completed only" and "SHIPMENT_COMPLETED_DT" in df.columns:
        df = df[_nonempty(df["SHIPMENT_COMPLETED_DT"])]
    if use_dates and d_from and d_to and "SHIPMENT_CREATED_DT" in df.columns:
        cr = pd.to_datetime(df["SHIPMENT_CREATED_DT"], errors="coerce")
        df = df[(cr >= pd.Timestamp(d_from)) & (cr < pd.Timestamp(d_to) + pd.Timedelta(days=1))]

    if len(df) == 0:
        st.warning("No rows after filtering — adjust the options in the sidebar.")
    elif not selected_milestones:
        st.warning("No milestones selected — pick at least one milestone in the sidebar.")
    else:
        # keep ALL_GROUPS journey order, restricted to what the user picked
        groups = [g for g in ALL_GROUPS if g[0] in selected_milestones]
        gnames = [g for g, _, _ in groups]

        overall, ent, fsub, N, pcols = compute(df, groups)
        st.subheader(f"Overall completeness — {N:,} shipments  ({len(gnames)} milestones)")
        ov = pd.DataFrame({"Milestone": gnames, "Completeness": [overall[g] for g in gnames]})
        st.dataframe(ov.style.format({"Completeness": "{:.1%}"}), use_container_width=True, hide_index=True)

        if include_raw:
            # df already reflects the current scope (All rows / Completed only) + date filter,
            # so the raw sheet name simply follows the scope selection — one sheet, not both.
            raw_sheet_name = "Raw Data - Completed" if scope == "Completed only" else "Raw Data - All"
            with st.spinner("Building workbook with raw-data sheet…"):
                xlsx = build_full_workbook(overall, ent, fsub, N, pcols, gnames, raw_df=df, raw_sheet_name=raw_sheet_name)
            st.caption(f"Workbook includes a '{raw_sheet_name}' tab ({len(df):,} rows).")
        else:
            xlsx = build_workbook(overall, ent, fsub, N, pcols, gnames)

        st.download_button("⬇️  Download completeness workbook (.xlsx)", data=xlsx,
                           file_name="OD2D_Completeness_By_Milestone.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.success("Workbook ready — click the button above to download.")
else:
    st.info("Upload a file to begin. Expected columns include the TS_* milestone timestamps, "
            "MASTER_FFW_NAME, MASTER_NVOCC_NAME, MASTER_CARRIER_NAME, and BOOKING_LEVEL.")
